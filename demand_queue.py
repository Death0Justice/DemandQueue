from collections import deque
from functools import partial
import sys, openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle
from datetime import datetime
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QIntValidator, QIcon, QCloseEvent, QKeySequence, QCursor
from PyQt5.QtCore import Qt
        
class DemandQueue(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.date_format = '%Y-%m-%d %H:%M'
        self.date_format_c = '%Y-%m-%d %H：%M'
        self.editing = False
        self.initShortcut()
        self.initHistory()
        self.initUI()
        
    def initShortcut(self):
        shortcut = QShortcut(QKeySequence('Ctrl+S'), self)
        shortcut.activated.connect(self.updateXLSX)
        
    def initHistory(self):
        self.queue:deque[list] = deque()
        try:
            h = openpyxl.load_workbook('history.xlsx')
        except:
            h = Workbook()
            h.active.title = '待完成点播'
            h.create_sheet('已完成点播')
            self.initHeaders(h, True)
            h.save('history.xlsx')
        
        for idx, row in enumerate(h['待完成点播'].values):
            if idx == 0:
                continue
            demand = list(row)
            if isinstance(demand[1], datetime):
                demand[1] = datetime.strftime(demand[1], self.date_format)
            self.queue.append(demand)
        self.history_sheet = h['已完成点播']
        
        self.history_book = h
    
    def initHeaders(self, h: Workbook, first=False):
        h['待完成点播'].append([
            "老板名称",
            "点播时间",
            "点播内容"
        ])
        h['待完成点播'].freeze_panes = 'A2'
        h['待完成点播'].row_dimensions[1].height = 20
        h['待完成点播'][1][0].alignment = Alignment(horizontal='center', vertical='center')
        h['待完成点播'][1][1].alignment = Alignment(horizontal='center', vertical='center')
        h['待完成点播'][1][2].alignment = Alignment(horizontal='center', vertical='center')
        h['待完成点播'].column_dimensions['A'].width = 25
        h['待完成点播'].column_dimensions['B'].width = 18
        h['待完成点播'].column_dimensions['C'].width = 80
        
        if first:
            # Because named styles only
            # needs to be added once
            date_style = NamedStyle(name='custom_datetime', number_format='YYYY-MM-DD HH:MM')
            h.add_named_style(date_style)
            
            # Because completed list is not truncated
            # after creation, do not append the headers
            h['已完成点播'].append([
                "老板名称",
                "点播时间",
                "点播内容",
                "完成时间"
            ])
            h['已完成点播'].freeze_panes = 'A2'
            h['已完成点播'].row_dimensions[1].height = 20
            h['已完成点播'][1][0].alignment = Alignment(horizontal='center', vertical='center')
            h['已完成点播'][1][1].alignment = Alignment(horizontal='center', vertical='center')
            h['已完成点播'][1][2].alignment = Alignment(horizontal='center', vertical='center')
            h['已完成点播'][1][3].alignment = Alignment(horizontal='center', vertical='center')
            h['已完成点播'].column_dimensions['A'].width = 25
            h['已完成点播'].column_dimensions['B'].width = 18
            h['已完成点播'].column_dimensions['C'].width = 80
            h['已完成点播'].column_dimensions['D'].width = 18
    
    def initUI(self):
            
        self.table = self.construct_table(self.queue)
        
        edit = QWidget(self)
        edit_layout = QHBoxLayout()
        
        name_comp = QWidget(edit)
        name_layout = QVBoxLayout()
        name4name = QLabel("老板名称", name_comp)
        self.name = QTextEdit(name_comp)
        name_layout.addWidget(name4name)
        name_layout.addWidget(self.name)
        name_comp.setLayout(name_layout)
        
        desc_comp = QWidget(edit)
        desc_layout = QVBoxLayout()
        name4desc = QLabel("点播内容", desc_comp)
        self.desc = QTextEdit(desc_comp)
        desc_layout.addWidget(name4desc)
        desc_layout.addWidget(self.desc)
        desc_comp.setLayout(desc_layout)
        
        date_comp = QWidget(edit)
        date_layout = QVBoxLayout()
        name4date = QLabel("点播时间", date_comp)
        self.date = QTextEdit(date_comp)
        date_layout.addWidget(name4date)
        date_layout.addWidget(self.date)
        date_comp.setLayout(date_layout)
        
        edit_layout.addWidget(name_comp)
        edit_layout.addWidget(date_comp)
        edit_layout.addWidget(desc_comp)
        edit.setLayout(edit_layout)
        
        insert = QWidget(self)
        insert_layout = QHBoxLayout()
        quick = QPushButton("快速添加", insert)
        quick.clicked.connect(self.quick_action)
        append = QPushButton("添加点播", insert)
        append.clicked.connect(self.append_queue)
        push = QPushButton("插播", insert)
        push.clicked.connect(self.push_queue)
        pop = QPushButton("完成点播", insert)
        pop.clicked.connect(self.pop_queue)
        insert_layout.addWidget(quick)
        insert_layout.addWidget(append)
        insert_layout.addWidget(push)
        insert_layout.addWidget(pop)
        insert.setLayout(insert_layout)
        
        advanced = QWidget(self)
        adv_layout = QHBoxLayout()
        sort_button = QPushButton("按时间排序", advanced)
        sort_button.clicked.connect(self.sort)
        delete_button = QPushButton("删除错误格式点播", advanced)
        delete_button.clicked.connect(self.del_unformatted)
        adv_layout.addStretch()
        adv_layout.addWidget(sort_button)
        adv_layout.addWidget(delete_button)
        adv_layout.addStretch()
        advanced.setLayout(adv_layout)
        
        self.layout = QVBoxLayout()
        self.layout.addWidget(self.table)
        self.layout.addWidget(edit, alignment=Qt.AlignBottom)
        self.layout.addWidget(insert, alignment=Qt.AlignBottom)
        self.layout.addWidget(advanced, alignment=Qt.AlignBottom)
        
        self.setGeometry(0, 0, 960, 960)
        self.setLayout(self.layout)
        
    def construct_table(self, queue: deque[list]) -> QTableWidget:
        table = QTableWidget(self)
        n = len(queue)
        table.setRowCount(n)
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["老板名称", "点播时间", "点播内容"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Interactive)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        table.setColumnWidth(1, 170)
        table.setContextMenuPolicy(Qt.CustomContextMenu)
        table.customContextMenuRequested.connect(self.contextMenu)
        for i in range(n):
            table.setItem(i, 0, QTableWidgetItem(queue[i][0]))
            table.setItem(i, 1, QTableWidgetItem(queue[i][1]))
            table.setItem(i, 2, QTableWidgetItem(queue[i][2]))
        table.itemChanged.connect(self.updateHistory)
        return table

    def contextMenu(self):
        pos = QCursor.pos()
        item_pos = self.table.viewport().mapFromGlobal(pos)
        item = self.table.itemAt(item_pos)
        # print(f'Dealing with item {item.text()}')
        menu = QMenu(self.table)
        edit = QAction("详情", menu)
        edit.triggered.connect(partial(self.showDemand, item))
        delete = QAction("删除点播", menu)
        delete.triggered.connect(partial(self.deleteDemand, item))
        complete = QAction("完成点播", menu)
        complete.triggered.connect(partial(self.completeDemand, item))
        insert = QAction("插入点播", menu)
        insert.triggered.connect(partial(self.insert_queue, item.row()))
        menu.addAction(edit)
        menu.addAction(delete)
        menu.addAction(complete)
        menu.addAction(insert)
        menu.exec_(pos)
    
    def deleteDemand(self, item: QTableWidgetItem):
        if len(self.queue) == 0:
            return
        row = item.row()
        del self.queue[row]
        self.table.removeRow(row)
        
        self.editing = True
    
    def completeDemand(self, item: QTableWidgetItem):
        if len(self.queue) == 0:
            return
        row = item.row()
        done = self.queue[row]
        done.append(datetime.now())
        if self.validate(done):
            done[1] = datetime.strptime(done[1], self.date_format)
        self.history_sheet.append(done)
        last = self.history_sheet.max_row
        self.history_sheet[last][1].style = 'custom_datetime'
        self.history_sheet[last][2].alignment = Alignment(wrap_text=True)
        self.history_sheet[last][3].style = 'custom_datetime'
        del self.queue[row]
        self.table.removeRow(row)
        
        self.editing = True
    
    def showDemand(self, item: QTableWidgetItem):
        if not item:
            return
        row = item.row()
        demand = self.queue[row]
        dialog = QDialog(self)
        dialog.setWindowTitle(f'点播详情')
        dialog.setWindowIcon(QIcon('isaac.ico'))
        dialog.setWindowFlag(Qt.WindowContextHelpButtonHint, False)
        dialog.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        dialog_layout = QVBoxLayout()
        dialog_layout.addWidget(QLabel(f'{demand[0]}的点播'), alignment=Qt.AlignCenter)
        info = QLabel(demand[2], dialog)
        info.setWordWrap(True)
        dialog_layout.addWidget(info, alignment=Qt.AlignCenter)
        dialog_layout.addWidget(QLabel(f'于{demand[1]}', dialog), alignment=Qt.AlignCenter)
        dialog.setLayout(dialog_layout)
        dialog.exec()
        
    def updateHistory(self, item: QTableWidgetItem):
        if not item:
            return
        row = item.row()
        col = item.column()
        self.queue[row][col] = item.text()
        self.editing = True

    def append_queue(self):
        self.insert_queue(self.table.rowCount())
        
    def push_queue(self):
        self.insert_queue(0, cut=True)
        
    def pop_queue(self):
        if len(self.queue) == 0:
            # Empty history
            # Can add pop-up here
            return
        done = self.queue.popleft()
        done.append(datetime.now())
        if self.validate(done):
            done[1] = datetime.strptime(done[1], self.date_format)
        self.history_sheet.append(done)
        last = self.history_sheet.max_row
        self.history_sheet[last][1].style = 'custom_datetime'
        self.history_sheet[last][2].alignment = Alignment(wrap_text=True)
        self.history_sheet[last][3].style = 'custom_datetime'
        self.table.removeRow(0)
        
        self.editing = True
    
    def del_unformatted(self):
        if len(self.queue) == 0:
            return
        while self.queue and not self.validate(self.queue[-1]):
            self.queue.pop()
            self.table.removeRow(self.table.rowCount() - 1)
        self.editing = True
    
    def insert_queue(self, row, cut=False):
        name = self.name.toPlainText()
        date = self.date.toPlainText()
        desc = self.desc.toPlainText()
        new = [name, date, desc]
        if cut:
            # Pushing
            new = [new[0], new[1], "插播: " + new[2]]
            self.queue.appendleft(new)
        else:
            # Appending or Inserting
            self.queue.insert(row, new)
        
        self.table.insertRow(row)
        self.table.setItem(row, 0, QTableWidgetItem(new[0]))
        self.table.setItem(row, 1, QTableWidgetItem(new[1]))
        self.table.setItem(row, 2, QTableWidgetItem(new[2]))
        
        self.editing = True
    
    def quick_action(self):
        # # Save the csv file before using buggy feature
        # self.updateXLSX()
        hint = "请输入以下格式点播：\n"
        hint += "[老板名称]\n"
        hint += "[点播时间]\n"
        hint += "[点播内容]\n"
        hint += "请用半角分号\";\"分隔多个点播\n"
        hint += "请注意：多个点播不支持插播"
        text, ok = QInputDialog.getMultiLineText(self, "快速输入", hint)

        if ok:
            table = str(text).split(";")
            if len(table) > 1:
                for line in table:
                    if not line or line.isspace():
                        continue
                    row = [s for s in line.split("\n") if s]
                    # Appending
                    name = row[0]
                    date = row[1]
                    desc = '\n'.join(row[2:])
                    self.queue.append([name, date, desc])
                    n = self.table.rowCount()
                    self.table.insertRow(n)
                    self.table.setItem(n, 0, QTableWidgetItem(name))
                    self.table.setItem(n, 1, QTableWidgetItem(date))
                    self.table.setItem(n, 2, QTableWidgetItem(desc))
            else:
                row = [s for s in str(text).split("\n") if s]
                name = row[0]
                date = row[1]
                desc = "\n".join(row[2:])
                self.name.setText(name)
                self.date.setText(date)
                self.desc.setText(desc)
        
        self.editing = True
    
    def sort(self):
        cut_queue = []
        normal_queue = []
        unformatted_queue = []
        for h in self.queue:
            if self.validate(h):
                if str(h[2]).startswith("插播"):
                    cut_queue.append(h)
                else:
                    normal_queue.append(h)
            else:
                unformatted_queue.append(h)
        cut_queue.sort(key=lambda h: datetime.strptime(h[1], self.date_format), reverse=True)
        normal_queue.sort(key=lambda h: datetime.strptime(h[1], self.date_format))
        
        self.queue = deque(cut_queue + normal_queue + unformatted_queue)
        # print(self.queue)
        self.table.clearContents()
        for i, h in enumerate(self.queue):
            self.table.setItem(i, 0, QTableWidgetItem(h[0]))
            self.table.setItem(i, 1, QTableWidgetItem(h[1]))
            self.table.setItem(i, 2, QTableWidgetItem(h[2]))
        
        self.editing = True
        
        conclusion = QDialog(self)
        conclusion.setWindowTitle("排序结果")
        conclusion.setWindowIcon(QIcon('isaac.ico'))
        conclusion.setWindowFlag(Qt.WindowContextHelpButtonHint, False)
        congrats = QLabel("排序完成！", conclusion)
        result = QLabel(f'发现{len(unformatted_queue)}个格式错误的点播。', conclusion)
        conclusion_layout = QVBoxLayout()
        conclusion_layout.addWidget(congrats, alignment=Qt.AlignCenter)
        conclusion_layout.addWidget(result, alignment=Qt.AlignCenter)
        conclusion.setLayout(conclusion_layout)
        conclusion.setMinimumSize(360, 160)
        conclusion.exec()
    
    def updateXLSX(self) -> bool:
        # print(self.queue)
        queue = self.history_book['待完成点播']
        self.history_book.remove(queue)
        self.history_book.create_sheet('待完成点播', 0)
        self.initHeaders(self.history_book)
        queue = self.history_book['待完成点播']
        for idx, row in enumerate(self.queue):
            queue.cell(idx + 2, 1).value = row[0]
            if self.validate(row):
                date = datetime.strptime(row[1], self.date_format)
                queue.cell(idx + 2, 2).value = date
            else:
                queue.cell(idx + 2, 2).value = row[1]
            queue.cell(idx + 2, 3).value = row[2]
            queue[idx + 2][1].style = 'custom_datetime'
            queue[idx + 2][2].alignment = Alignment(wrap_text=True)
        try:
            self.history_book.save('history.xlsx')
            self.editing = False
            return True
        except PermissionError:
            not_permitted = Popup('保存失败！', '保存失败，请查看是否有其他程序正在占用文档。请关闭占用的程序后再点击重试。', ["重试", "取消"], QIcon('isaac.ico'))
            reply = not_permitted.do()
            match reply:
                case '重试':
                    return self.updateXLSX()
                case '取消':
                    return False
    
    def validate(self, demand: list) -> bool:
        try:
            time = datetime.strptime(demand[1], self.date_format)
            demand[1] = time.strftime(self.date_format)
            return True
        except ValueError:
            # print('En date match failed')
            pass
        
        try:
            time = datetime.strptime(demand[1], self.date_format_c)
            demand[1] = time.strftime(self.date_format)
            return True
        except ValueError:
            # print('Cn date match failed')
            pass
        
        return ''

    def closeEvent(self, e: QCloseEvent):
        if not self.editing:
            # Already saved or not editted at all
            e.accept()
        else:
            # Unsaved edit content
            confirm = Popup('保存','是否保存当前点播队列？', ['是', '否', '取消'], QIcon('isaac.ico'))
            reply = confirm.do()
            match reply:
                case '是':
                    if self.updateXLSX():
                        e.accept()
                    else:
                        e.ignore()
                case '否':
                    e.accept()
                case '取消':
                    e.ignore()

class Popup(QMessageBox):
    def __init__(
            self, 
            title, 
            text, 
            buttons = ["Ok"],
            icon = QIcon()
        ):
        
        super(Popup, self).__init__()
        self.setWindowTitle(title)
        self.setText(text)
        self.setWindowIcon(icon)
        self.buttons = buttons
        for txt in self.buttons:
            b = QPushButton(txt)
            self.addButton(b, QMessageBox.NoRole)
            
    def do(self):
        answer = self.exec_()
        text = self.buttons[answer]
        return text
        
        
if __name__ == "__main__":
    qApp = QApplication(sys.argv)
    dq = DemandQueue()
    dq.setWindowTitle("点播队列")
    dq.setWindowIcon(QIcon('isaac.ico'))
    dq.setFixedSize(960, 960)
    # Center the window
    qr = dq.frameGeometry()
    cp = QDesktopWidget().availableGeometry().center()
    qr.moveCenter(cp)
    dq.move(qr.topLeft())
    
    dq.show()
    
    sys.exit(qApp.exec_())